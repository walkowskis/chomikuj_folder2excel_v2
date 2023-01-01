from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
import time
import openpyxl
import logging
import datetime
import data

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

c_handler = logging.StreamHandler()
c_handler.setLevel(logging.DEBUG)
c_format = logging.Formatter('%(asctime)s - %(levelname)s: %(message)s', datefmt='%H:%M:%S')
c_handler.setFormatter(c_format)
logger.addHandler(c_handler)

f_handler = logging.FileHandler('data_log.log')
f_handler.setLevel(logging.WARNING)
f_format = logging.Formatter('%(asctime)s - %(levelname)s: %(message)s', datefmt='%d-%m-%Y %H:%M:%S')
f_handler.setFormatter(f_format)
logger.addHandler(f_handler)

today = datetime.date.today()
year = '2022' # today.year

logger.info('Getting Started...')
webdriver_service = Service(data.webdriver_path)
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('--window-size=1920,1080')
driver = webdriver.Chrome(service=webdriver_service, options=options)
driver.implicitly_wait(10)

driver.get(data.website_url + str(year))
driver.find_element(By.XPATH, "//button[@type='submit']").click()
driver.find_element(By.ID, 'topBarLogin').send_keys(data.login)
driver.find_element(By.ID, 'topBarPassword').send_keys(data.pwd)
driver.find_element(By.ID, 'topBarPassword').send_keys(Keys.ENTER)
logger.info('Logged successfully.')
time.sleep(5)
driver.find_element(By.ID, 'Password').send_keys(data.folder_owner_pwd)
driver.find_element(By.ID, 'Password').send_keys(Keys.ENTER)
logger.info('Successfully logged into the folder.')
time.sleep(5)
folder_list = [i.text for i in driver.find_elements(By.XPATH, '//div[@id="foldersList"]//td')]
logger.info('Recently added folder: %s', folder_list[-1:][0])
logger.info('Other folders for the current year: %s', ', '.join(folder_list[:-1]))
selected_folder = input('Choose the folder to scan: ')

driver.get(data.website_url + str(year) + '/' + selected_folder)


def next_page(driver):
    element = driver.find_elements(By.XPATH, '//div[contains(@class,"filename txt")]//a[contains(@class,'
                                             '"expanderHeader downloadAction")]')
    if element:
        return element
    else:
        return False


logger.info('Starting the scan:')
item_list = []
hyperlinks = []
while True:
    item = WebDriverWait(driver, 10).until(next_page)
    logger.info('   - %s items added;', len(item))
    for x in item:
        item_list.append(x.text)
        hyperlinks.append(x.get_attribute('href'))
    try:
        driver.find_element(By.XPATH, '//a[@class="right"]').click()
    except:
        logger.info('Folder scan completed - %s items added.', len(item_list))
        break

    time.sleep(5)

logger.info('Opening the xlsx file.')
wb = openpyxl.load_workbook('ebook.xlsx')
sheet = wb.create_sheet(selected_folder, 0)
sheet.append(["Author", "Title", "Folder", "Rating", "Votes", "Category", "Date of the first edition", "Date of the "
                                                                                                       "first PL "
                                                                                                       "edition",
              "Pages No", "Tags", "Description"])

counter = 0
list_clear = []
for item in item_list:
    var = (item[:-5]).rsplit(' - ', 1)
    if len(var) == 2:
        var.append(hyperlinks[counter])
        counter += 1
        list_clear.append(var)
    else:
        counter += 1
        logger.warning('Unrecognized author or title in item: %s', var[:])

logger.info('Saving the list to file.')
i = len(sheet['A'])
for title, author, url in list_clear:
    i += 1
    sheet.cell(row=i, column=1).value = author
    sheet.cell(row=i, column=2).value = title
    sheet.cell(row=i, column=3).hyperlink = url
    sheet.cell(row=i, column=3).value = selected_folder
    sheet.cell(row=i, column=3).style = "Hyperlink"
    logger.debug('  - the %s line is written;', i)

logger.info('Saving the file.')
i = 0
while i < 3:
    try:
        wb.save('ebook.xlsx')
        break
    except PermissionError:
        i += 1
        print(f"You Don't Have Permission to Access the File, retry in 15 sec.")
        time.sleep(15)
